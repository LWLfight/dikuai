#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import json
import requests
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException, WebDriverException
import time
import re
import random
import openpyxl
from openpyxl.styles import Alignment
from urllib.parse import urljoin


def download_pdf(pdf_url, save_dir="应急管理部PDF文件"):
    """
    下载PDF文件
    
    参数:
        pdf_url: PDF文件的URL
        save_dir: 保存目录
    
    返回:
        str: 下载的文件名，如果失败返回None
    """
    try:
        # 创建保存目录
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
        
        # 从URL中提取文件名
        filename = os.path.basename(pdf_url.split('?')[0])  # 去除URL参数
        
        # 如果文件名无效，使用时间戳生成
        if not filename or not filename.endswith('.pdf'):
            timestamp = int(time.time())
            filename = f"document_{timestamp}.pdf"
        
        filepath = os.path.join(save_dir, filename)
        
        print(f"      📥 开始下载PDF: {filename}")
        
        # 下载PDF文件
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        response = requests.get(pdf_url, headers=headers, timeout=30)
        response.raise_for_status()
        
        # 保存文件
        with open(filepath, 'wb') as f:
            f.write(response.content)
        
        file_size = os.path.getsize(filepath)
        print(f"      ✓ PDF下载成功: {filename} ({file_size/1024:.1f} KB)")
        
        return filename
        
    except Exception as e:
        print(f"      ✗ PDF下载失败: {e}")
        return None


def is_pdf_link(url):
    """
    判断URL是否为PDF文件链接
    
    参数:
        url: URL地址
    
    返回:
        bool: True表示是PDF链接
    """
    if not url:
        return False
    
    # 检查URL是否以.pdf结尾（忽略大小写和URL参数）
    url_lower = url.lower().split('?')[0]
    return url_lower.endswith('.pdf')


def try_extract_content_from_current_page(driver):
    """
    尝试从当前页面提取正文内容
    
    返回:
        tuple: (content, is_valid) 
        - content: 提取的正文内容
        - is_valid: 是否为有效正文(True/False)
    """
    print(f"      📋 开始尝试提取正文...")
    try:
        # 尝试多种选择器(按优先级从高到低排列)
        content_selectors = [
            'div.Custom_UnionStyle',      # 最内层正文容器(优先级最高)
            'div.TRS_Editor',              # TRS编辑器内容区
            'div#zoom',                     # ID定位(唯一性强)
            'div.editor-content',           # 编辑器内容
            'div.zhenwen_neir',             # 正文内容外层
            'div.content',                  # 通用内容容器
            'div.article-content'           # 文章内容容器
        ]
        
        content_div = None
        used_selector = None
        for selector in content_selectors:
            try:
                # 直接查找元素,不使用WebDriverWait(页面已加载完成)
                print(f"      尝试选择器: {selector}")
                content_div = driver.find_element(By.CSS_SELECTOR, selector)
                used_selector = selector
                print(f"      ✓ 找到元素: {selector}")
                break
            except Exception as e:
                continue
        
        if content_div:
            print(f"      📝 开始提取段落文本...")
            # 提取所有p标签的文本内容
            paragraphs = content_div.find_elements(By.TAG_NAME, 'p')
            print(f"      找到 {len(paragraphs)} 个p标签")
            content_parts = []
            for p in paragraphs:
                text = p.text
                if text and text.strip():  # 确保不是纯空白
                    content_parts.append(text)
            
            detail_content = '\n'.join(content_parts)
            print(f"      提取到 {len(content_parts)} 个有效段落,总字符数: {len(detail_content)}")
            
            # 判断是否为有效正文(至少50个字符,且不只是链接或空内容)
            if len(detail_content) > 50:
                print(f"      ✓ 提取成功,使用选择器: {used_selector}, 正文字数: {len(detail_content)}")
                return detail_content, True
            else:
                print(f"      ⚠ 提取内容过少({len(detail_content)}字符),视为无效正文")
                return "", False
        else:
            print(f"      ⚠ 未找到正文容器")
            return "", False
            
    except Exception as e:
        print(f"      ⚠ 提取正文异常: {type(e).__name__}: {e}")
        import traceback
        traceback.print_exc()
        return "", False


def extract_links_from_related_section(driver):
    """
    从"相关链接"区域提取链接列表
    
    返回:
        list: 链接URL列表
    """
    try:
        related_links = []
        
        # 查找"相关链接"区域
        # 根据HTML结构: <div class="xglj zw-file2"> 或包含"相关链接"文本的区域
        related_sections = driver.find_elements(By.XPATH, '//div[contains(@class, "xglj")] | //h2[contains(text(), "相关链接")]/..')
        
        if not related_sections:
            print(f"      ⚠ 未找到相关链接区域")
            return []
        
        print(f"      找到 {len(related_sections)} 个相关链接区域")
        
        for section in related_sections:
            try:
                # 在该区域内查找所有<li>下的<a>标签(排除标题链接)
                list_items = section.find_elements(By.CSS_SELECTOR, 'li a, ul li a')
                
                for link in list_items:
                    href = link.get_attribute('href')
                    text = link.text.strip()
                    
                    if href and text:
                        # 排除PDF链接、下载链接等
                        if not href.endswith('.pdf') and not href.endswith('.doc') and not href.endswith('.docx'):
                            # 排除"相关链接"标题本身
                            if text != '相关链接' and len(text) > 0:
                                related_links.append({
                                    'url': href,
                                    'text': text
                                })
                                print(f"      发现相关链接: {text}")
            except Exception as e:
                print(f"      处理相关链接区域时出错: {e}")
                continue
        
        return related_links
        
    except Exception as e:
        print(f"      查找相关链接失败: {e}")
        return []


def extract_detail_content_recursive_with_source(driver, max_depth=4, current_depth=0):
    """
    递归提取详情页正文内容,支持链接跳转,并返回正文来源URL
    
    参数:
        driver: Selenium WebDriver实例
        max_depth: 最大递归深度,防止无限循环(默认4层)
        current_depth: 当前递归深度
    
    返回:
        tuple: (content, source_url)
        - content: 提取的正文内容,如果无法提取则返回空字符串
        - source_url: 最终提取到正文的页面URL
    """
    if current_depth >= max_depth:
        print(f"      ⚠ 达到最大递归深度 {max_depth},停止提取")
        return "", driver.current_url
    
    indent = "      " + "  " * current_depth  # 缩进用于日志显示
    
    print(f"{indent}🔍 第{current_depth + 1}层页面分析...")
    
    # 步骤1: 尝试提取当前页面的正文
    content, is_valid = try_extract_content_from_current_page(driver)
    
    if is_valid:
        # 有有效正文,直接返回
        current_url = driver.current_url
        print(f"{indent}✅ 第{current_depth + 1}层找到有效正文 (来源: {current_url})")
        return content, current_url
    
    # 步骤2: 没有有效正文,查找相关链接
    print(f"{indent}🔗 当前页面无有效正文,查找相关链接...")
    related_links = extract_links_from_related_section(driver)
    
    if not related_links:
        print(f"{indent}❌ 无相关链接,无法提取正文")
        return "", driver.current_url
    
    # 步骤3: 依次打开相关链接,递归提取
    print(f"{indent}📎 发现 {len(related_links)} 个相关链接,开始逐个尝试...")
    
    all_contents = []
    all_sources = []
    original_url = driver.current_url  # 保存原始页面URL
    
    for idx, link_info in enumerate(related_links):
        link_url = link_info['url']
        link_text = link_info['text']
        
        print(f"{indent}➡️  [{idx + 1}/{len(related_links)}] 尝试链接: {link_text}")
        
        try:
            # 打开链接
            driver.get(link_url)
            time.sleep(2)  # 缩短等待时间
            
            # 递归提取正文
            sub_content, sub_source_url = extract_detail_content_recursive_with_source(
                driver, 
                max_depth=max_depth, 
                current_depth=current_depth + 1
            )
            
            if sub_content:
                print(f"{indent}✅ 从链接 '{link_text}' 提取到正文 ({len(sub_content)}字符)")
                all_contents.append(sub_content)
                all_sources.append(sub_source_url)
            else:
                print(f"{indent}⚠️ 链接 '{link_text}' 未提取到正文")
            
            # 返回原始页面,继续处理下一个链接
            driver.get(original_url)
            time.sleep(1)  # 缩短等待时间
            
        except Exception as e:
            print(f"{indent}❌ 处理链接 '{link_text}' 时出错: {e}")
            # 尝试返回原始页面
            try:
                driver.get(original_url)
                time.sleep(1)
            except:
                pass
            continue
    
    # 合并所有内容
    if all_contents:
        combined_content = '\n\n'.join(all_contents)
        # 如果有多个来源,用第一个作为主要来源
        primary_source = all_sources[0] if all_sources else original_url
        print(f"{indent}📝 共从 {len(all_contents)} 个链接提取到正文,总计 {len(combined_content)} 字符")
        return combined_content, primary_source
    else:
        print(f"{indent}❌ 所有链接均未提取到有效正文")
        return "", original_url


def open_website():
    """
    使用Selenium打开应急管理部官网
    """
    driver = None
    # 目标网址 - 应急管理部特别重大事故调查报告页面
    target_url = "https://www.mem.gov.cn/gk/sgcc/tbzdsgdcbg/index.shtml"
    
    # 配置浏览器选项
    chrome_options = webdriver.ChromeOptions()
    
    # 性能优化配置
    chrome_options.add_argument('--no-sandbox')              # 提升启动稳定性与速度
    chrome_options.add_argument('--disable-dev-shm-usage')   # 解决资源限制
    chrome_options.add_argument('--disable-extensions')      # 禁用扩展,减少初始化开销
    chrome_options.add_argument('--disable-plugins')         # 禁用插件,减少初始化开销
    chrome_options.add_argument('--disable-images')          # 禁用图片加载,大幅提升页面加载速度
    # 注意: 不禁用GPU加速,保留GPU以提升渲染速度
    chrome_options.add_argument('--log-level=3')             # 减少日志输出，提升速度
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')  # 隐藏自动化特征
    chrome_options.add_argument('--disable-infobars')        # 禁用信息栏
    chrome_options.add_argument('--window-size=1280,800')    # 设置较小窗口尺寸,减少渲染开销

    try:
        print("正在初始化浏览器...")
        start_time = time.time()
        
        # 使用本地缓存的ChromeDriver
        chromedriver_path = os.path.join(
            os.environ['USERPROFILE'],
            '.wdm', 'drivers', 'chromedriver', 'win64',
            '147.0.7727.56', 'chromedriver-win32', 'chromedriver.exe'
        )
        
        if os.path.exists(chromedriver_path):
            print(f"  使用本地ChromeDriver: {chromedriver_path}")
            service = Service(chromedriver_path)
            driver = webdriver.Chrome(service=service, options=chrome_options)
            print("✓ 浏览器初始化成功")
        else:
            print(f"  ✗ 未找到ChromeDriver: {chromedriver_path}")
            raise FileNotFoundError(f"ChromeDriver不存在: {chromedriver_path}")
        
        init_time = time.time() - start_time
        print(f"✓ 浏览器初始化完成，耗时: {init_time:.2f}秒")
        
        # 设置隐式等待
        driver.implicitly_wait(5)
        
        print(f"正在访问: {target_url}")
        
        # 打开目标网址
        page_start = time.time()
        driver.get(target_url)
        
        # 显示等待：等待页面标题加载完成
        try:
            WebDriverWait(driver, 10).until(
                lambda d: d.title is not None and len(d.title) > 0
            )
            page_load_time = time.time() - page_start
            print(f"✓ 页面标题: {driver.title}")
            print(f"✓ 页面加载耗时: {page_load_time:.2f}秒")
        except TimeoutException:
            print("⚠ 警告：页面标题加载超时，但页面可能已打开")
        
        # 获取当前URL
        current_url = driver.current_url
        print(f"当前实际URL: {current_url}")
        
        # 打印页面源代码长度
        page_source_len = len(driver.page_source)
        print(f"页面源代码长度: {page_source_len} 字符")
        
        if page_source_len > 1000:
            print("✓ 页面加载成功！")
        else:
            print("⚠ 警告：页面内容较少，可能未完全加载")
        
    except WebDriverException as e:
        print(f"✗ 浏览器驱动错误: {e}")
        print("\n可能的解决方案:")
        print("1. 检查Chrome浏览器是否正常安装")
        print("2. 确认ChromeDriver版本与Chrome浏览器版本匹配")
        print(f"3. 检查ChromeDriver路径是否正确: {chromedriver_path if 'chromedriver_path' in locals() else '未知'}")
        return None
    except Exception as e:
        print(f"✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        return None
    
    return driver


def check_anti_crawl(driver, consecutive_failures=0):
    """
    检测是否触发反爬机制,并根据失败次数执行相应的等待策略
    
    参数:
        driver: Selenium WebDriver实例
        consecutive_failures: 连续失败次数
    
    返回:
        bool: True表示检测到反爬,需要等待; False表示正常
    """
    anti_crawl_indicators = [
        lambda: "captcha" in driver.current_url.lower(),
        lambda: "verify" in driver.current_url.lower(),
        lambda: "blocked" in driver.current_url.lower(),
        lambda: "forbidden" in driver.current_url.lower(),
        lambda: len(driver.page_source) < 1000,
        lambda: len(driver.find_elements(By.XPATH, "//div[contains(text(), '验证')]")) > 0,
        lambda: len(driver.find_elements(By.XPATH, "//div[contains(text(), 'captcha')]")) > 0,
    ]
    
    is_blocked = any(indicator() for indicator in anti_crawl_indicators)
    
    if is_blocked:
        if consecutive_failures == 0:
            wait_time = random.uniform(10, 15)
            level = "第1级"
        elif consecutive_failures == 1:
            wait_time = random.uniform(20, 25)
            level = "第2级"
        elif consecutive_failures == 2:
            wait_time = random.uniform(50, 60)
            level = "第3级"
        else:
            wait_time = random.uniform(100, 120)
            level = "第4级(最高)"
        
        print(f"\n⚠️ 检测到反爬机制触发! ({level})")
        print(f"   连续失败次数: {consecutive_failures + 1}")
        print(f"   等待时间: {wait_time:.1f} 秒")
        print(f"   正在等待...")
        
        time.sleep(wait_time)
        return True
    
    return False


def handle_anti_crawl_recovery(driver, consecutive_failures):
    """
    反爬恢复策略:刷新页面或重新导航
    """
    try:
        print(f"   尝试恢复访问...")
        current_url = driver.current_url
        
        driver.refresh()
        time.sleep(5)
        
        if len(driver.page_source) > 1000:
            print(f"   ✓ 页面刷新成功,已恢复正常访问")
            return True
        else:
            print(f"   刷新无效,尝试重新导航...")
            driver.get(current_url)
            time.sleep(5)
            
            if len(driver.page_source) > 1000:
                print(f"   ✓ 重新导航成功,已恢复正常访问")
                return True
            else:
                print(f"   ✗ 恢复失败")
                return False
                
    except Exception as e:
        print(f"   ✗ 恢复过程出错: {e}")
        return False


def save_single_record_to_excel(record, output_filename="应急管理部爬取结果.xlsx"):
    """
    将单条数据追加保存到Excel文件
    
    参数:
        record: 包含title, url, date, content, content_source_url, pdf_filename的字典
        output_filename: 输出文件名
    """
    try:
        # 如果文件不存在，创建新文件并写入表头
        if not os.path.exists(output_filename):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "事故调查报告"
            
            # 设置表头(新增"正文来源链接"和"pdf"列)
            headers = ["标题", "链接", "时间", "正文", "正文来源链接", "pdf"]
            ws.append(headers)
            
            # 设置表头样式
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置列宽
            ws.column_dimensions['A'].width = 60  # 标题
            ws.column_dimensions['B'].width = 80  # 链接
            ws.column_dimensions['C'].width = 15  # 时间
            ws.column_dimensions['D'].width = 100  # 正文
            ws.column_dimensions['E'].width = 80  # 正文来源链接
            ws.column_dimensions['F'].width = 40  # pdf
        else:
            # 文件已存在，加载现有文件
            wb = openpyxl.load_workbook(output_filename)
            ws = wb.active
        
        # 添加新数据行(新增content_source_url和pdf_filename字段)
        ws.append([
            record.get('title', ''),
            record.get('url', ''),
            record.get('date', ''),
            record.get('content', ''),
            record.get('content_source_url', ''),  # 正文来源链接
            record.get('pdf_filename', '')  # PDF文件名
        ])
        
        # 获取当前行号（最后一行）
        current_row = ws.max_row
        
        # 设置新行的格式
        for cell in ws[current_row]:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[current_row].height = 14
        
        # 保存文件（带重试机制）
        max_retries = 3
        for attempt in range(max_retries):
            try:
                wb.save(output_filename)
                return True
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(2)
                else:
                    print(f"  ✗ 保存失败: 文件可能被Excel打开，请关闭后重试")
                    raise
            except Exception as e:
                print(f"  ✗ 保存失败: {e}")
                raise
                
    except Exception as e:
        print(f"保存到Excel时出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def append_single_record_to_txt(record, output_filename="应急管理部爬取结果.txt", index=None):
    """
    将单条数据追加保存到TXT文件
    
    参数:
        record: 包含title, url, date, content, content_source_url, pdf_filename的字典
        output_filename: 输出文件名
        index: 可选，数据的序号
    """
    try:
        # 如果没有提供索引，尝试通过文件现有内容估算
        if index is None:
            index = 1
            if os.path.exists(output_filename):
                try:
                    with open(output_filename, 'r', encoding='utf-8') as count_f:
                        lines = sum(1 for _ in count_f)
                        index = (lines // 9) + 1
                except:
                    pass

        with open(output_filename, 'a', encoding='utf-8') as f:
            f.write(f"{'='*80}\n")
            f.write(f"[{index}] {record.get('title', '')}\n")
            f.write(f"{'='*80}\n")
            f.write(f"链接: {record.get('url', '')}\n")
            f.write(f"时间: {record.get('date', '')}\n")
            
            # 如果有正文来源URL，也记录下来
            source_url = record.get('content_source_url')
            if source_url:
                f.write(f"正文来源: {source_url}\n")
            
            # 如果有PDF文件，记录PDF文件名
            pdf_filename = record.get('pdf_filename')
            if pdf_filename:
                f.write(f"PDF文件: {pdf_filename}\n")
            
            f.write(f"\n正文:\n")
            f.write(f"{record.get('content', '')}\n")
            f.write(f"\n\n")
        
        print(f"  ✓ 已提取并保存第 {index} 条数据")
        return True
                
    except Exception as e:
        print(f"保存到TXT时出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def save_to_txt(data_list, output_filename="应急管理部爬取结果.txt"):
    """
    将数据保存到TXT文件
    
    参数:
        data_list: 数据列表，每个元素是包含title, url, date, content, content_source_url, pdf_filename的字典
        output_filename: 输出文件名
    """
    if not data_list:
        print("没有数据可保存")
        return False

    try:
        with open(output_filename, 'w', encoding='utf-8') as f:
            for idx, item in enumerate(data_list, 1):
                f.write(f"{'='*80}\n")
                f.write(f"[{idx}] {item.get('title', '')}\n")
                f.write(f"{'='*80}\n")
                f.write(f"链接: {item.get('url', '')}\n")
                f.write(f"时间: {item.get('date', '')}\n")
                # 如果有正文来源URL,也记录下来
                if item.get('content_source_url'):
                    f.write(f"正文来源: {item.get('content_source_url', '')}\n")
                # 如果有PDF文件，记录PDF文件名
                if item.get('pdf_filename'):
                    f.write(f"PDF文件: {item.get('pdf_filename', '')}\n")
                f.write(f"\n正文:\n")
                f.write(f"{item.get('content', '')}\n")
                f.write(f"\n\n")
        
        print(f"\n✓ 数据已保存到TXT文件: {output_filename}")
        print(f"  共保存 {len(data_list)} 条数据")
        return True
                
    except Exception as e:
        print(f"保存到TXT时出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def save_to_excel(data_list, output_filename="应急管理部爬取结果.xlsx"):
    """
    将数据保存到Excel文件
    
    参数:
        data_list: 数据列表，每个元素是包含title, url, date, content, content_source_url, pdf_filename的字典
        output_filename: 输出文件名
    """
    if not data_list:
        print("没有数据可保存")
        return False

    try:
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "事故调查报告"
        
        # 设置表头(新增"正文来源链接"和"pdf"列)
        headers = ["标题", "链接", "时间", "正文", "正文来源链接", "pdf"]
        ws.append(headers)
        
        # 设置表头样式
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加数据
        for item in data_list:
            ws.append([
                item.get('title', ''),
                item.get('url', ''),
                item.get('date', ''),
                item.get('content', ''),
                item.get('content_source_url', ''),  # 正文来源链接
                item.get('pdf_filename', '')  # PDF文件名
            ])
        
        # 设置列宽
        ws.column_dimensions['A'].width = 60  # 标题
        ws.column_dimensions['B'].width = 80  # 链接
        ws.column_dimensions['C'].width = 15  # 时间
        ws.column_dimensions['D'].width = 100  # 正文
        ws.column_dimensions['E'].width = 80  # 正文来源链接
        ws.column_dimensions['F'].width = 40  # pdf
        
        # 设置所有行的行高为14，并开启自动换行
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row[0].row].height = 14
        
        # 保存文件（带重试机制）
        max_retries = 3
        for attempt in range(max_retries):
            try:
                wb.save(output_filename)
                print(f"\n✓ 数据已保存到Excel文件: {output_filename}")
                print(f"  共保存 {len(data_list)} 条数据")
                return True
            except PermissionError:
                if attempt < max_retries - 1:
                    print(f"  文件被占用，3秒后重试...")
                    time.sleep(3)
                else:
                    print(f"  ✗ 保存失败: 文件可能被Excel打开，请关闭后重试")
                    raise
            except Exception as e:
                print(f"  ✗ 保存失败: {e}")
                raise
                
    except Exception as e:
        print(f"保存到Excel时出错: {e}")
        import traceback
        traceback.print_exc()
        return False


def extract_page_results(driver, all_data, output_filename="应急管理部爬取结果.xlsx"):
    """
    提取当前页面的所有事故报告链接,并进入详情页获取内容
    
    参数:
        driver: Selenium WebDriver实例
        all_data: 用于累积数据的列表
        output_filename: 输出文件名
    """
    seen_urls = set()
    consecutive_failures = 0
    
    try:
        print(f"\n{'='*60}")
        print(f"开始提取当前页面的事故报告...")
        print(f"{'='*60}")
        
        # 检测反爬机制
        if check_anti_crawl(driver, consecutive_failures):
            if not handle_anti_crawl_recovery(driver, consecutive_failures):
                print(f"   恢复失败,跳过当前页")
                return all_data
            else:
                consecutive_failures = 0
        
        # 查找页面上所有的table元素（每个table包含最多5个结果）
        tables = driver.find_elements(By.XPATH, '//table[@width="790"]')
        
        if not tables:
            print("未找到包含结果的table元素")
            return all_data
        
        print(f"找到 {len(tables)} 个表格区域")
        
        result_index = len(all_data)
        
        # 遍历每个table
        for table_idx in range(len(tables)):
            print(f"\n处理第 {table_idx + 1} 个表格...")
            
            # 重新获取当前table（避免stale element）
            try:
                current_tables = driver.find_elements(By.XPATH, '//table[@width="790"]')
                if table_idx >= len(current_tables):
                    print(f"  表格索引超出范围")
                    break
                table = current_tables[table_idx]
            except Exception as e:
                print(f"  重新获取表格失败: {e}")
                continue
            
            # 在当前table中查找所有的<a>标签
            try:
                links = table.find_elements(By.TAG_NAME, 'a')
            except Exception as e:
                print(f"  查找链接失败: {e}")
                continue
            
            if not links:
                print(f"  第 {table_idx + 1} 个表格中没有找到链接")
                continue
            
            print(f"  找到 {len(links)} 个链接")
            
            # 遍历每个链接
            for link_idx in range(len(links)):
                try:
                    # 重新获取当前链接（避免stale element）
                    current_tables = driver.find_elements(By.XPATH, '//table[@width="790"]')
                    if table_idx >= len(current_tables):
                        break
                    current_links = current_tables[table_idx].find_elements(By.TAG_NAME, 'a')
                    if link_idx >= len(current_links):
                        break
                    link = current_links[link_idx]
                    
                    # 提取标题和日期
                    title_full = link.text.strip()
                    href = link.get_attribute('href')
                    
                    if not title_full or not href:
                        continue
                    
                    # 分离标题和日期
                    date_match = re.search(r'(\d{4}-\d{2}-\d{2})$', title_full)
                    if date_match:
                        date_str = date_match.group(1)
                        title_text = title_full[:date_match.start()].strip()
                    else:
                        date_str = ""
                        title_text = title_full
                    
                    # 构建完整URL
                    if href.startswith('../../'):
                        base_url = "https://www.mem.gov.cn/gk/sgcc/tbzdsgdcbg/"
                        full_url = base_url + href[6:]
                    elif href.startswith('./'):
                        base_url = "https://www.mem.gov.cn/gk/sgcc/tbzdsgdcbg/"
                        full_url = base_url + href[2:]
                    elif href.startswith('/'):
                        full_url = "https://www.mem.gov.cn" + href
                    else:
                        full_url = href
                    
                    # 去重检查
                    if full_url in seen_urls:
                        print(f"  跳过重复: {title_text}")
                        continue
                    
                    result_index += 1
                    print(f"\n  [{result_index}] {title_text}")
                    print(f"      日期: {date_str}")
                    print(f"      URL: {full_url}")
                    
                    # 保存当前页面URL以便返回
                    list_page_url = driver.current_url
                    
                    # 检查是否为PDF链接
                    pdf_filename = None
                    detail_content = ""
                    content_source_url = full_url
                    
                    if is_pdf_link(full_url):
                        # 是PDF链接，直接下载
                        print(f"      📄 检测到PDF链接，开始下载...")
                        pdf_filename = download_pdf(full_url)
                        
                        if pdf_filename:
                            detail_content = f"[PDF文件已下载: {pdf_filename}]"
                        else:
                            detail_content = "[PDF文件下载失败]"
                            
                        result_info = {
                            'title': title_text,
                            'url': full_url,
                            'date': date_str,
                            'content': detail_content,
                            'content_source_url': full_url,
                            'pdf_filename': pdf_filename if pdf_filename else ''
                        }
                    else:
                        # 不是PDF链接，正常访问详情页
                        driver.get(full_url)
                        time.sleep(2)
                        
                        # 检测详情页是否触发反爬
                        if check_anti_crawl(driver, consecutive_failures):
                            if not handle_anti_crawl_recovery(driver, consecutive_failures):
                                print(f"      ⚠ 详情页触发反爬,跳过此条数据")
                                consecutive_failures += 1
                                driver.get(list_page_url)
                                time.sleep(2)
                                continue
                            else:
                                consecutive_failures = 0
                        
                        time.sleep(1)
                        
                        # 获取详情页信息
                        detail_url = driver.current_url
                        detail_title = driver.title
                        print(f"      详情页标题: {detail_title}")
                        
                        # 使用递归函数提取详情页正文内容
                        detail_content, content_source_url = extract_detail_content_recursive_with_source(driver, max_depth=4)
                        
                        if not detail_content:
                            print(f"      ⚠ 未能提取到有效正文")
                            detail_content = "无法提取正文内容"
                            content_source_url = detail_url
                        
                        print(f"      📍 正文来源: {content_source_url}")
                        
                        result_info = {
                            'title': title_text,
                            'url': detail_url,
                            'date': date_str,
                            'content': detail_content,
                            'content_source_url': content_source_url,
                            'pdf_filename': ''  # 非PDF链接，留空
                        }
                    
                    all_data.append(result_info)
                    seen_urls.add(full_url)
                    
                    # 立即保存到Excel和TXT文件
                    print(f"      正在保存数据...")
                    txt_filename = output_filename.replace('.xlsx', '.txt')
                    
                    excel_ok = save_single_record_to_excel(result_info, output_filename)
                    txt_ok = append_single_record_to_txt(result_info, txt_filename, index=result_index)
                    
                    if excel_ok and txt_ok:
                        print(f"      ✓ 已提取并保存第 {result_index} 条数据 (Excel + TXT)")
                    else:
                        print(f"      ✗ 保存失败，但数据已添加到内存列表")
                    
                    # 如果不是PDF链接，需要返回列表页
                    if not is_pdf_link(full_url):
                        driver.get(list_page_url)
                        time.sleep(2)
                    
                    # 关键词间的冷却时间
                    cool_down = random.uniform(2, 5)
                    print(f"      冷却 {cool_down:.1f} 秒...")
                    time.sleep(cool_down)
                    
                except Exception as e:
                    print(f"      处理链接时出错: {e}")
                    # 尝试返回列表页
                    try:
                        if 'list_page_url' in locals():
                            driver.get(list_page_url)
                            time.sleep(2)
                    except:
                        pass
                    continue
        
        print(f"\n{'='*60}")
        print(f"当前页面提取完成! 共提取 {result_index - len(all_data) + len(all_data)} 条数据")
        print(f"{'='*60}")
        
        return all_data
        
    except Exception as e:
        print(f"提取过程中发生错误: {e}")
        import traceback
        traceback.print_exc()
        return all_data


def navigate_to_next_page(driver, current_page_num):
    """
    通过直接构造URL跳转到下一页
    
    参数:
        driver: Selenium WebDriver实例
        current_page_num: 当前页码(从1开始)
    
    返回:
        tuple: (success, next_url)
        - success: True表示成功跳转, False表示已是最后一页或失败
        - next_url: 下一页的URL
    """
    try:
        print(f"\n准备跳转到第 {current_page_num + 1} 页...")
        
        # 策略1: 尝试从JS变量获取总页数
        try:
            total_pages = driver.execute_script("return countPage;")
            if total_pages is not None and total_pages > 0:
                print(f"  ✓ 从JS变量获取总页数: {total_pages}")
                
                # 检查是否已到最后一页
                if current_page_num >= total_pages:
                    print(f"  ⚠ 已是最后一页(共{total_pages}页)")
                    return False, None
            else:
                total_pages = None
        except Exception as e:
            print(f"  ⚠ 无法从JS获取总页数: {e}")
            total_pages = None
        
        # 构造下一页URL
        base_url = "https://www.mem.gov.cn/gk/sgcc/tbzdsgdcbg/"
        if current_page_num == 1:
            # 当前是第1页,要跳转到第2页,第2页的URL是 index_1.shtml
            next_url = base_url + "index_1.shtml"
        else:
            # 当前是第N页,要跳转到第N+1页
            # 第N+1页的URL是 index_N.shtml
            next_url = base_url + f"index_{current_page_num}.shtml"
        
        print(f"  目标URL: {next_url}")
        
        # 保存当前URL用于对比
        current_url = driver.current_url
        
        # 直接导航到下一页URL
        driver.get(next_url)
        time.sleep(3)  # 等待页面加载
        
        # 验证是否成功跳转
        new_url = driver.current_url
        if new_url != current_url:
            print(f"  ✓ 跳转成功")
            print(f"  实际URL: {new_url}")
            
            # 尝试验证页码
            try:
                current_page_js = driver.execute_script("return currentPage;")
                if current_page_js is not None:
                    actual_page = current_page_js + 1  # JS中currentPage从0开始
                    print(f"  ✓ 当前页码验证: 第{actual_page}页")
            except:
                pass
            
            return True, new_url
        else:
            print(f"  ✗ URL未变化,跳转可能失败")
            return False, None
            
    except Exception as e:
        print(f"  ✗ 跳转失败: {e}")
        import traceback
        traceback.print_exc()
        return False, None


def crawl_all_pages(driver, output_filename="应急管理部爬取结果.xlsx"):
    """
    爬取所有页面的数据
    
    参数:
        driver: Selenium WebDriver实例
        output_filename: 输出文件名
    """
    all_data = []
    page_num = 1
    max_pages = 100  # 最大页数限制,防止无限循环
    
    print(f"\n{'='*80}")
    print(f"开始爬取应急管理部事故报告")
    print(f"目标URL: https://www.mem.gov.cn/gk/sgcc/tbzdsgdcbg/index.shtml")
    print(f"输出文件: {output_filename}")
    print(f"{'='*80}")
    
    # 清空TXT文件(准备实时追加)
    txt_filename = output_filename.replace('.xlsx', '.txt')
    try:
        with open(txt_filename, 'w', encoding='utf-8') as f:
            pass  # 创建空文件
        print(f"✓ 已初始化TXT文件: {txt_filename}")
    except Exception as e:
        print(f"⚠ 初始化TXT文件失败: {e}")
    
    # 尝试从JS变量获取总页数
    total_pages = None
    try:
        total_pages = driver.execute_script("return countPage;")
        if total_pages is not None and total_pages > 0:
            print(f"\n✓ 从JS变量获取总页数: {total_pages}")
        else:
            print(f"\n⚠ 无法从JS获取有效总页数,将使用max_pages限制({max_pages}页)")
            total_pages = None
    except Exception as e:
        print(f"\n⚠ 无法从JS获取总页数: {e},将使用max_pages限制({max_pages}页)")
        total_pages = None
    
    while page_num <= max_pages:
        # 如果从JS获取到总页数,则使用它作为上限
        if total_pages is not None and page_num > total_pages:
            print(f"\n{'='*80}")
            print(f"已到达最后一页(共{total_pages}页)")
            print(f"{'='*80}")
            break
        
        print(f"\n{'#'*80}")
        print(f"# 正在处理第 {page_num} 页")
        print(f"{'#'*80}")
        
        # 保存当前页面HTML用于调试
        try:
            debug_filename = f"mem_page_{page_num}.html"
            with open(debug_filename, "w", encoding="utf-8") as f:
                f.write(driver.page_source)
            print(f"已保存第 {page_num} 页HTML到 {debug_filename}")
        except:
            pass
        
        # 提取当前页面的结果
        all_data = extract_page_results(driver, all_data, output_filename)
        
        if not all_data or len(all_data) == 0:
            print(f"\n第 {page_num} 页没有提取到数据")
        
        # 尝试跳转到下一页
        success, next_url = navigate_to_next_page(driver, page_num)
        
        if not success:
            print(f"\n{'='*80}")
            print(f"已到达最后一页或无法继续翻页")
            print(f"{'='*80}")
            break
        
        page_num += 1
        
        # 页面间的冷却时间
        cool_down = random.uniform(3, 6)
        print(f"\n页面间冷却 {cool_down:.1f} 秒...")
        time.sleep(cool_down)
    
    # 最后统一保存到Excel和TXT
    print(f"\n{'='*80}")
    print(f"爬取完成! 共爬取 {page_num} 页,总计 {len(all_data)} 条数据")
    print(f"正在保存到文件...")
    
    # 生成TXT文件名(将.xlsx替换为.txt)
    txt_filename = output_filename.replace('.xlsx', '.txt')
    
    # 保存Excel
    save_to_excel(all_data, output_filename)
    
    # 保存TXT
    save_to_txt(all_data, txt_filename)
    
    print(f"{'='*80}")
    
    return all_data


def main():
    """
    主函数
    """
    driver = None
    try:
        # 打开网站
        driver = open_website()
        
        if not driver:
            print("浏览器初始化失败,程序退出")
            return
        
        # 等待页面完全加载
        time.sleep(3)
        
        # 爬取所有页面
        results = crawl_all_pages(driver, output_filename="应急管理部爬取结果.xlsx")
        
        print(f"\n程序执行完成! 共获取 {len(results)} 条数据")
        
    except KeyboardInterrupt:
        print("\n\n用户中断程序")
    except Exception as e:
        print(f"\n程序执行出错: {e}")
        import traceback
        traceback.print_exc()
    finally:
        # 关闭浏览器
        if driver:
            try:
                print("\n正在关闭浏览器...")
                driver.quit()
                print("浏览器已关闭")
            except:
                pass


if __name__ == "__main__":
    main()
